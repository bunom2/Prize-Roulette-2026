import logging
import asyncio
import uuid
import os
import random
import signal
import sys
import io
from datetime import datetime

# Библиотеки для Google
import gspread
from oauth2client.service_account import ServiceAccountCredentials

# Библиотеки для Yandex/Excel
import requests
import openpyxl

import aiohttp
from aiohttp import web
from aiogram import Bot, Dispatcher, types
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.utils import executor
from aiogram.utils.exceptions import MessageNotModified, TerminatedByOtherGetUpdates
from dotenv import load_dotenv

# --- КОНФИГУРАЦИЯ ---
load_dotenv()
BOT_TOKEN = os.getenv("BOT_TOKEN")
ADMIN_IDS = list(map(int, os.getenv("ADMIN_IDS", "").split(",")))

# Выбор источника данных: 'google' или 'yandex'
DATA_SOURCE = os.getenv("DATA_SOURCE", "google").lower()

# Настройки Google
GOOGLE_SHEET_ID = os.getenv("SHEET_ID")

# Настройки Яндекс
YANDEX_TOKEN = os.getenv("YANDEX_TOKEN")
YANDEX_FILE_PATH = "roulette.xlsx" # Имя файла в корне Яндекс.Диска

logging.basicConfig(level=logging.INFO)

# Глобальная блокировка для операций с данными (чтобы не было гонки потоков)
db_lock = asyncio.Lock()

# --- КЛАСС ДЛЯ РАБОТЫ С ДАННЫМИ (АБСТРАКЦИЯ) ---

class DataManager:
    """Управляет данными в зависимости от выбранного источника (Google или Yandex)."""
    
    @staticmethod
    def get_google_client():
        scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
        filename = "service_account.json"
        if not os.path.exists(filename):
            logging.error(f"Файл {filename} не найден! Проверьте Secret Files на Render.")
        creds = ServiceAccountCredentials.from_json_keyfile_name(filename, scope)
        return gspread.authorize(creds)

    # --- ОБЩИЕ МЕТОДЫ (ВЫЗЫВАЮТ СПЕЦИФИЧЕСКИЕ) ---

    @staticmethod
    async def get_prizes():
        async with db_lock: # Блокируем доступ для других запросов, пока читаем
            if DATA_SOURCE == 'google':
                return DataManager._get_prizes_google()
            elif DATA_SOURCE == 'yandex':
                return await DataManager._get_prizes_yandex()

    @staticmethod
    async def record_winner(user, prize):
        async with db_lock:
            if DATA_SOURCE == 'google':
                DataManager._record_winner_google(user, prize)
            elif DATA_SOURCE == 'yandex':
                await DataManager._record_winner_yandex(user, prize)

    @staticmethod
    async def add_tokens(tokens):
        async with db_lock:
            if DATA_SOURCE == 'google':
                DataManager._add_tokens_google(tokens)
            elif DATA_SOURCE == 'yandex':
                await DataManager._add_tokens_yandex(tokens)

    @staticmethod
    async def check_token(token):
        # Google читает быстро, блокировка может быть излишней, но для надежности оставим
        # Для Яндекса блокировка критична (внутри методов)
        if DATA_SOURCE == 'google':
            return DataManager._check_token_google(token)
        elif DATA_SOURCE == 'yandex':
            return await DataManager._check_token_yandex(token)

    @staticmethod
    async def mark_token_used(token_data):
        async with db_lock:
            if DATA_SOURCE == 'google':
                DataManager._mark_token_used_google(token_data)
            elif DATA_SOURCE == 'yandex':
                await DataManager._mark_token_used_yandex(token_data)

    # --- GOOGLE IMPLEMENTATION ---
    @staticmethod
    def _get_prizes_google():
        client = DataManager.get_google_client()
        sheet = client.open_by_key(GOOGLE_SHEET_ID).worksheet("Prizes")
        all_records = sheet.get_all_records()
        available = []
        for idx, item in enumerate(all_records, start=2):
            try:
                limit = int(item['Лимит'])
                issued = int(item['Выдано'])
                if limit - issued > 0:
                    item['row_idx'] = idx
                    available.append(item)
            except ValueError: continue
        return available

    @staticmethod
    def _record_winner_google(user, prize):
        client = DataManager.get_google_client()
        sh = client.open_by_key(GOOGLE_SHEET_ID)
        # Обновляем счетчик призов
        sh.worksheet("Prizes").update_cell(prize['row_idx'], 4, int(prize['Выдано']) + 1)
        # Пишем победителя
        row = [
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"), 
            str(user.id), 
            f"@{user.username}" if user.username else "NoUsername", 
            prize['Название приза']
        ]
        sh.worksheet("Winners").append_row(row)

    @staticmethod
    def _add_tokens_google(tokens):
        client = DataManager.get_google_client()
        # Добавляем сразу пачкой
        client.open_by_key(GOOGLE_SHEET_ID).worksheet("Tokens").append_rows([[t, 'active'] for t in tokens])

    @staticmethod
    def _check_token_google(token):
        client = DataManager.get_google_client()
        ws = client.open_by_key(GOOGLE_SHEET_ID).worksheet("Tokens")
        try:
            cell = ws.find(token)
            if cell:
                # Возвращаем статус, ряд, колонку статуса (она следующая за токеном)
                return ws.cell(cell.row, cell.col + 1).value, cell.row, cell.col + 1
        except: pass
        return None, None, None

    @staticmethod
    def _mark_token_used_google(token_data):
        # token_data = (status, row, col)
        row, col = token_data[1], token_data[2]
        if row and col:
            client = DataManager.get_google_client()
            client.open_by_key(GOOGLE_SHEET_ID).worksheet("Tokens").update_cell(row, col, 'used')

    # --- YANDEX / EXCEL IMPLEMENTATION ---
    
    @staticmethod
    def _yandex_headers():
        return {'Authorization': f'OAuth {YANDEX_TOKEN}'}

    @staticmethod
    def _download_excel():
        """Скачивает файл с Я.Диска в память."""
        url = "https://cloud-api.yandex.net/v1/disk/resources/download"
        params = {'path': YANDEX_FILE_PATH}
        # Получаем ссылку на скачивание
        resp = requests.get(url, headers=DataManager._yandex_headers(), params=params)
        if resp.status_code != 200:
            logging.error(f"Yandex Download Error: {resp.text}")
            raise Exception(f"Ошибка скачивания: {resp.status_code}")
        
        download_url = resp.json()['href']
        file_resp = requests.get(download_url)
        return io.BytesIO(file_resp.content)

    @staticmethod
    def _upload_excel(buffer):
        """Загружает файл из памяти на Я.Диск (перезапись)."""
        url = "https://cloud-api.yandex.net/v1/disk/resources/upload"
        params = {'path': YANDEX_FILE_PATH, 'overwrite': 'true'}
        
        # Получаем ссылку на загрузку
        resp = requests.get(url, headers=DataManager._yandex_headers(), params=params)
        if resp.status_code != 200:
            logging.error(f"Yandex Upload Link Error: {resp.text}")
            raise Exception("Не удалось получить ссылку для загрузки")
            
        upload_url = resp.json()['href']
        buffer.seek(0)
        requests.put(upload_url, files={'file': buffer})

    @staticmethod
    async def _get_prizes_yandex():
        try:
            # Скачиваем файл в память
            wb = openpyxl.load_workbook(DataManager._download_excel())
            ws = wb['Prizes']
            prizes = []
            # Предполагаем структуру: A=ID, B=Name, C=Limit, D=Issued
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # row[0]=ID, row[1]=Name, row[2]=Limit, row[3]=Issued
                if len(row) >= 3 and row[1] and row[2] is not None:
                    issued = row[3] if len(row) > 3 and row[3] is not None else 0
                    if int(row[2]) - int(issued) > 0:
                        prizes.append({
                            'Название приза': row[1],
                            'Лимит': row[2],
                            'Выдано': issued,
                            'row_idx': row_idx
                        })
            return prizes
        except Exception as e:
            logging.error(f"Yandex Read Prizes Error: {e}")
            return []

    @staticmethod
    async def _record_winner_yandex(user, prize):
        wb = openpyxl.load_workbook(DataManager._download_excel())
        
        # 1. Обновляем Prizes
        ws_prizes = wb['Prizes']
        # cell(row, column). D=4
        curr_val = ws_prizes.cell(row=prize['row_idx'], column=4).value
        curr_val = int(curr_val) if curr_val else 0
        ws_prizes.cell(row=prize['row_idx'], column=4).value = curr_val + 1
        
        # 2. Пишем в Winners
        if 'Winners' not in wb.sheetnames:
            wb.create_sheet('Winners')
        ws_winners = wb['Winners']
        ws_winners.append([
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            str(user.id),
            f"@{user.username}" if user.username else "NoUsername",
            prize['Название приза']
        ])
        
        # Сохраняем и загружаем
        buffer = io.BytesIO()
        wb.save(buffer)
        DataManager._upload_excel(buffer)

    @staticmethod
    async def _add_tokens_yandex(tokens):
        wb = openpyxl.load_workbook(DataManager._download_excel())
        if 'Tokens' not in wb.sheetnames:
            wb.create_sheet('Tokens')
        ws = wb['Tokens']
        for t in tokens:
            ws.append([t, 'active'])
        
        buffer = io.BytesIO()
        wb.save(buffer)
        DataManager._upload_excel(buffer)

    @staticmethod
    async def _check_token_yandex(token):
        # В Яндексе мы блокируем чтение тоже, чтобы не прочитать старый файл, пока другой поток пишет
        async with db_lock:
            wb = openpyxl.load_workbook(DataManager._download_excel())
            if 'Tokens' not in wb.sheetnames:
                return None, None, None
            ws = wb['Tokens']
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
                if row[0] == token:
                    # Возвращаем статус, row_idx, и сам объект token (не используется)
                    status = row[1] if len(row) > 1 else None
                    return status, row_idx, None
            return None, None, None

    @staticmethod
    async def _mark_token_used_yandex(token_data):
        # token_data = (status, row_idx, _)
        row_idx = token_data[1]
        
        wb = openpyxl.load_workbook(DataManager._download_excel())
        ws = wb['Tokens']
        # Статус в колонке B (2)
        ws.cell(row=row_idx, column=2).value = 'used'
        
        buffer = io.BytesIO()
        wb.save(buffer)
        DataManager._upload_excel(buffer)


# --- WEB SERVER & KEEP ALIVE ---
web_runner = None
async def health_check(request):
    return web.Response(text=f"Bot running. Source: {DATA_SOURCE.upper()}")

async def start_web_server():
    global web_runner
    app = web.Application()
    app.router.add_get('/', health_check)
    web_runner = web.AppRunner(app)
    await web_runner.setup()
    port = int(os.getenv("PORT", 8080))
    site = web.TCPSite(web_runner, '0.0.0.0', port)
    await site.start()
    logging.info(f"Web server started on port {port}. Source: {DATA_SOURCE}")

async def keep_alive():
    url = os.getenv("RENDER_EXTERNAL_URL")
    if not url: return
    logging.info(f"Запущен Keep-Alive для {url}")
    while True:
        await asyncio.sleep(9 * 60)
        try:
            async with aiohttp.ClientSession() as session:
                async with session.get(url) as resp:
                    logging.info(f"Keep-alive status: {resp.status}")
        except: pass

# --- БОТ ИНИЦИАЛИЗАЦИЯ ---
bot = Bot(token=BOT_TOKEN)
dp = Dispatcher(bot, storage=MemoryStorage())

@dp.message_handler(commands=['generate'], user_id=ADMIN_IDS)
async def cmd_generate(message: types.Message):
    try:
        count = int(message.get_args())
    except:
        await message.reply("Использование: /generate <N>")
        return
    
    new_tokens = [str(uuid.uuid4())[:8] for _ in range(count)]
    try:
        msg = await message.reply(f"Пишу в {DATA_SOURCE.upper()}... Ждите.")
        await DataManager.add_tokens(new_tokens)
    except Exception as e:
        logging.error(f"DB Error: {e}")
        await message.reply("Ошибка записи базы данных (проверьте логи).")
        return

    bot_username = (await bot.get_me()).username
    lines = [f"https://t.me/{bot_username}?start={t}" for t in new_tokens]
    
    with open("links.txt", "w") as f: f.write("\n".join(lines))
    
    caption = f"Готово: {count} шт. ({DATA_SOURCE.upper()})"
    await message.reply_document(open("links.txt", "rb"), caption=caption)
    os.remove("links.txt")

@dp.message_handler(commands=['start'])
async def cmd_start(message: types.Message):
    token = message.get_args()
    if not token:
        await message.answer("Для участия нужна ссылка.")
        return

    status, row, col = await DataManager.check_token(token)
    
    if status == 'active':
        markup = types.InlineKeyboardMarkup()
        markup.add(types.InlineKeyboardButton("🚀 Запустить систему", callback_data=f"step1:{token}"))
        await message.answer("👋 Привет! Система готова.", reply_markup=markup)
    elif status == 'used':
        await message.answer("Эта ссылка уже была использована.")
    else:
        await message.answer("Неверная ссылка (или ошибка доступа к базе).")

@dp.callback_query_handler(lambda c: c.data.startswith('step1:'))
async def process_step_1(c: types.CallbackQuery):
    token = c.data.split(":")[1]
    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("⚡ Зарядить на удачу ⚡", callback_data=f"step2:{token}"))
    await c.message.edit_text("📡 Связь с космосом установлена...\n🔄 Калибровка удачи... [████░░]\n🔎 Поиск лучших призов...", reply_markup=markup)

@dp.callback_query_handler(lambda c: c.data.startswith('step2:'))
async def process_step_2(c: types.CallbackQuery):
    token = c.data.split(":")[1]
    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("🎰 КРУТИТЬ РУЛЕТКУ! 🎰", callback_data=f"spin:{token}"))
    await c.message.edit_text("🔋 Энергия: 100%\n🍀 Удача: МАКСИМУМ\n🔥 Система готова к выдаче приза!", reply_markup=markup)

@dp.callback_query_handler(lambda c: c.data.startswith('spin:'))
async def process_spin(c: types.CallbackQuery):
    token = c.data.split(":")[1]
    
    # ПРОВЕРКА
    status, row, col = await DataManager.check_token(token)
    if status != 'active':
        await c.answer("Неактивно.")
        await c.message.delete()
        return

    try: await c.message.edit_reply_markup(reply_markup=None)
    except MessageNotModified: pass
    
    await bot.send_dice(c.from_user.id, emoji='🎰')
    await asyncio.sleep(2.5)
    
    try:
        # ВЫБОР И ЗАПИСЬ
        prizes = await DataManager.get_prizes()
        if not prizes:
             await bot.send_message(c.from_user.id, "Призы закончились! 😔")
             if row: await DataManager.mark_token_used((status, row, col))
             return

        won_prize = random.choice(prizes)
        
        # ЗАПИСЬ ПОБЕДИТЕЛЯ
        await DataManager.record_winner(c.from_user, won_prize)
        
        # ГАШЕНИЕ ТОКЕНА
        if row: await DataManager.mark_token_used((status, row, col))
        
        # ВАУ-ЭФФЕКТ
        await bot.send_message(
            c.from_user.id, 
            f"🎇🎇🎇 <b>БА-БАХ! ЕСТЬ КОНТАКТ!</b> 🎇🎇🎇\n\n"
            f"🎁 Твой приз: <b>{won_prize['Название приза']}</b>\n\n"
            f"🥳 Поздравляем с победой!", 
            parse_mode="HTML"
        )
    except Exception as e:
        logging.error(f"Error in spin: {e}")
        await bot.send_message(c.from_user.id, "Произошла техническая ошибка.")

async def on_startup(dp):
    # Обработка сигналов остановки
    def handle_signal(sig, frame):
        logging.warning(f"Получен сигнал {sig}. Останавливаемся...")
        asyncio.create_task(on_shutdown(dp))
        
    if sys.platform != "win32":
        signal.signal(signal.SIGTERM, handle_signal)
        signal.signal(signal.SIGINT, handle_signal)

    asyncio.create_task(keep_alive())
    await start_web_server()
    await bot.delete_webhook(drop_pending_updates=True)
    
    # Пауза для безопасного деплоя (убийства зомби)
    logging.info("⏳ Пауза 40 сек перед стартом Polling (Safe Deploy)...")
    await asyncio.sleep(40)
    logging.info("🚀 Старт Polling!")

async def on_shutdown(dp):
    logging.warning('Shutting down bot...')
    if web_runner: await web_runner.cleanup()
    await bot.close()
    await dp.storage.close()
    await dp.storage.wait_closed()
    logging.warning('Bot stopped completely.')

if __name__ == '__main__':
    try:
        executor.start_polling(dp, skip_updates=True, on_startup=on_startup, on_shutdown=on_shutdown)
    except TerminatedByOtherGetUpdates:
        logging.error("Конфликт обновлений. Перезапуск...")
        sys.exit(1)
        